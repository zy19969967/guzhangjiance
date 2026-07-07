import os
import numpy as np
import pandas as pd
from scipy.io import loadmat
from openpyxl import load_workbook


ROBOT_SENSOR_COLUMNS = [f"AI1-{idx:02d}" for idx in range(1, 19)]


def _normalize_robot_signal(signal):
    signal = signal.astype(np.float32, copy=False)
    minimum = signal.min(axis=0, keepdims=True)
    maximum = signal.max(axis=0, keepdims=True)
    scaled = (signal - minimum) / (maximum - minimum + 1e-8)
    return 2 * scaled - 1


def CWRU(item_path):
    axis = ["_DE_time", "_FE_time", "_BA_time"]
    datanumber = os.path.basename(item_path).split(".")[0]
    if eval(datanumber) < 100:
        realaxis = "X0" + datanumber + axis[0]
    else:
        realaxis = "X" + datanumber + axis[0]
    signal = loadmat(item_path)[realaxis]

    return signal


def MFPT(item_path):
    f = item_path.split("/")[-2]
    if f == 'normal':
        signal = (loadmat(item_path)["bearing"][0][0][1])
    else:
        signal = (loadmat(item_path)["bearing"][0][0][2])

    return signal


def PU(item_path):
    name = os.path.basename(item_path).split(".")[0]
    fl = loadmat(item_path)[name]
    signal = fl[0][0][2][0][6][2]  #Take out the data
    signal = signal.reshape(-1,1)

    return signal


def XJTU(item_path):
    fl = pd.read_csv(item_path)
    signal = fl["Horizontal_vibration_signals"]
    signal = signal.values.reshape(-1,1)

    return signal


def IMS(item_path):
    channel = {'normal': 0,
               'inner': 4,
               'outer': 0,
               'ball': 6}
    f = item_path.split("/")[-2]
    signal = np.loadtxt(item_path)[:, channel[f]]

    return signal


def JNU(item_path):
    fl = pd.read_csv(item_path)
    signal = fl.values
    
    return signal


def ROBOT(item_path):
    cache_path = None
    if not item_path.lower().endswith(".xlsx"):
        cache_path = item_path + ".file_norm.npz"
        if os.path.exists(cache_path):
            return np.load(cache_path)["signal"]
        with open(item_path, "r", encoding="utf-8") as f:
            item_path = f.readline().strip()

    workbook = load_workbook(item_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        worksheet.reset_dimensions()
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows)
        index_by_name = {str(value): idx for idx, value in enumerate(header)}
        missing = [column for column in ROBOT_SENSOR_COLUMNS if column not in index_by_name]
        if missing:
            raise ValueError(f"{item_path} is missing sensor columns: {missing}")

        sensor_indexes = [index_by_name[column] for column in ROBOT_SENSOR_COLUMNS]
        signal_rows = []
        for row in rows:
            try:
                signal_rows.append([float(row[idx]) for idx in sensor_indexes])
            except (TypeError, ValueError):
                continue
    finally:
        workbook.close()

    if not signal_rows:
        raise ValueError(f"{item_path} has no valid robot acceleration rows")
    signal = _normalize_robot_signal(np.asarray(signal_rows, dtype=np.float32))
    if cache_path is not None:
        np.savez_compressed(cache_path, signal=signal)
    return signal
